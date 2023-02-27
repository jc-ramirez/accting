import openpyxl
import datetime
import string

from openpyxl import Workbook
from openpyxl import load_workbook
mth = datetime.datetime.now().strftime("%b")
path = "9027.xlsx"

#def open_workbook(path):
  #wb = load_workbook(path)
  #print(f"Worksheet names: {wb.sheetnames}")
  #sheet = wb.active
  #print(sheet)

def firstmonth(wb, mth):
  if mth in wb.sheetnames:
    return False
  else:
    return True
  

def ld_rents(path, sheet):
  #wb = load_workbook(path)
  ws = wb[sheet]
  mth = datetime.datetime.now().strftime("%b")
  source = ws
  target = wb.copy_worksheet(source)
  target.title = mth
  #wss = wb["Rents Copy"]
  #print(wss.title)
  #wss.title = mth
  #wb["Feb"] = wb.copy_worksheet(ws)
  print(f"Worksheet names: {wb.sheetnames}")


def coord_conv(str):
  if ":" not in str:
    print("Please recheck coordinates.")
    return 0
  else:
    x = str.split(":")
    d = {"min_row" : int(x[0][1]), "min_col" : string.ascii_lowercase.index(x[0][0].lower()) + 1, "max_row" : int(x[1][1]) + 1, 
    "max_col" : string.ascii_lowercase.index(x[1][0].lower()) + 1}
    return d


def cp_values(sheet_name, cell_range):
  try:
    sh = wb[sheet_name]
    cdinates = coord_conv(cell_range)
    d = {}

    for i in range(cdinates['min_row'], cdinates['max_row']):
      cell_value_aptno = sh.cell(i,cdinates['min_col']).value
      cell_value_rent = sh.cell(i, cdinates['max_col']).value
      d[cell_value_aptno] = cell_value_rent

    for e in d:
      print(f"{e} = {d[e]}")
	
  except:
    print(f"'{sheet_name}' not found. Quitting.")
    return

  return d


def wr_values(dict, sheet_name, cell_range):
  try:
    sh = wb[sheet_name]
    cdinates = coord_conv(cell_range)
    ind = 0

    for i in range(cdinates['min_row'], cdinates['max_row']):
        sh.cell(i,cdinates['min_col']).value = list(dict.keys())[ind]
        sh.cell(i, cdinates['max_col']).value = list(dict.values())[ind]
        ind += 1

    for e in dict:
      print(f"{e} = {dict[e]}")
  
  except:
    print(f"'Couldn't write.")
    return

  return d

def create_workbook(path):
  wb = Workbook()
  mth = datetime.datetime.now().strftime("%b")
  #dttm = datetime.datetime.strptime(mylist[i], "%m/%d/%Y")
  sheet = wb.active
  sheet.cell(row=1,column=1).value = mth
  wb.save(path)
  
if __name__ == "__main__":
  wb = load_workbook(path)
  #ld_rents(wb, "Rents")
  if firstmonth(wb, mth):
    st_month = input("Esta es la primera vez que inicias sesión este mes. Te gustaria empezar la cuenta mensual?\nEscribi 'si' o 'no'\n")

    if st_month.lower() == 'si':
      #ts_month = ld_rents(path, "Rents")
      d = cp_values("Rents", "A2:B5")
      ws = wb.create_sheet(f"{mth}")
      print(f"Worksheet names: {wb.sheetnames}")
      wr_values(d, f"{mth}", "A2:B5")

  else:
    st_month = input("Con que te puedeo ayudar?\n")








Footer
© 2023 GitHub, Inc.
Footer navigation
Terms
Privacy
Security
Status
Docs
Contact GitHub
Pricing
API
Training
Blog
About
